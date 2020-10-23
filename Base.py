import faker
from openpyxl import Workbook

data = faker.Faker()
wb = Workbook()  
sheet = wb.active  

def displayWelcomeMessage():
    print("************Test data genarator************")

def enterNumberOfRecords():
    r = input("Please enter number of Records: ")
    return r

def selectOptions():
    print("Please select data which you want to generate: ")
    print("Enter 1 for full name")
    print("Enter 2 for email")
    print("Enter 3 for address")
    op = input("Please enter you choice(use comma in case of multiple option)-- ")
    return op

def GenerateDate(rec, data1):
    r = int(rec)
    li = data1.split(",")

    for i in range(1, r+1):
        count = 1
        for j in (li):
            if(j == '1'):
                sheet.cell(i, count).value = data.name()
                count = count +1
            elif(j == '2'):
                sheet.cell(i, count).value = data.email()
                count = count +1
            elif(j == '3'):
                sheet.cell(i, count).value = data.address()
                count = count +1
    
    wb.save("Result.xlsx")